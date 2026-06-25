"""Inventory connector: parse the Excel file into the inventory table.

Reads the most-recently-modified .xlsx/.xlsm in data/documents/inventory/, maps columns
heuristically (SKU, name, quantity, unit, location), and upserts rows. The cursor is the
file content hash, so an unchanged file is skipped — but quantities are re-read whenever
the file changes (the user updates it constantly).

Column detection is best-effort; if the office sheet uses unusual headers, override the
mapping later via Settings.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

from ..config import get_settings
from ..kv import get_setting

_settings = get_settings()
_HOSTFS = "/hostfs"  # the macOS filesystem is bind-mounted read-only here (see docker-compose)

_NAME_KEYS = ("name", "description", "item", "product")
_SKU_KEYS = ("sku", "part", "code", "item #", "item#")
_QTY_KEYS = ("quantity", "qty", "on hand", "onhand", "stock", "count")
_UNIT_KEYS = ("unit", "uom")
_LOC_KEYS = ("location", "loc ", "bin", "shelf", "warehouse")


def _match(header: str, keys: tuple[str, ...]) -> bool:
    h = (header or "").strip().lower()
    return any(k in h for k in keys)


def _latest_file() -> Path | None:
    folder = _settings.documents_dir / "inventory"
    files = [p for p in folder.glob("*") if p.suffix.lower() in (".xlsx", ".xlsm")]
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


def _configured_file() -> Path | None:
    """A path set in the admin UI: a file, OR a folder whose newest .xlsx is used (handy for
    date-stamped files like 'Master Inventory 4-6-26.xlsx'). The admin enters the real macOS
    path; we read it via the read-only host mount at /hostfs."""
    p = (get_setting("inventory_path") or "").strip()
    if not p:
        return None
    cp = Path(_HOSTFS + p) if p.startswith("/") else Path(p)
    if cp.is_dir():
        files = [x for x in cp.glob("*") if x.suffix.lower() in (".xlsx", ".xlsm")]
        return max(files, key=lambda x: x.stat().st_mtime) if files else None
    return cp


def _resolve_file() -> Path | None:
    # A path configured in the admin UI wins; otherwise use a file uploaded under Documents.
    return _configured_file() or _latest_file()


def sync(creds: dict, cursor: str | None) -> dict:
    path = _resolve_file()
    if not path:
        return {"inventory": [], "cursor": cursor}
    if not path.exists():
        raise ValueError(f"inventory file not found (looked at {path}); check the path is correct and under /Users")

    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    if digest == cursor:
        return {"inventory": [], "cursor": cursor}  # unchanged since last sync

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = (get_setting("inventory_sheet") or "").strip()
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inventory": [], "cursor": digest}

    header = [str(c) if c is not None else "" for c in rows[0]]
    col: dict[str, int] = {}
    for i, h in enumerate(header):
        if "name" not in col and _match(h, _NAME_KEYS):
            col["name"] = i
        if "sku" not in col and _match(h, _SKU_KEYS):
            col["sku"] = i
        if "quantity" not in col and _match(h, _QTY_KEYS):
            col["quantity"] = i
        if "unit" not in col and _match(h, _UNIT_KEYS):
            col["unit"] = i
        if "location" not in col and _match(h, _LOC_KEYS):
            col["location"] = i

    def cell(row, key):
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    records = []
    for row in rows[1:]:
        name, sku = cell(row, "name"), cell(row, "sku")
        if name is None and sku is None:
            continue
        qty = cell(row, "quantity")
        try:
            qty = float(qty) if qty is not None and str(qty).strip() != "" else None
        except (TypeError, ValueError):
            qty = None
        records.append(
            {
                "name": str(name) if name is not None else str(sku),
                "sku": str(sku) if sku is not None else None,
                "quantity": qty,
                "unit": str(cell(row, "unit")) if cell(row, "unit") is not None else None,
                "location": str(cell(row, "location"))
                if cell(row, "location") is not None
                else None,
                # JSON-safe snapshot of the source row.
                "raw": {h: ("" if v is None else str(v)) for h, v in zip(header, row)},
            }
        )

    return {"inventory": records, "cursor": digest}
